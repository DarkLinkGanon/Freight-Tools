from flask import Flask, request, render_template, send_file, make_response, jsonify
import pdfplumber
from openpyxl import Workbook, load_workbook
from io import BytesIO, StringIO
import os
import re
import tempfile
import base64
import csv
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime

app = Flask(__name__)

FUEL_LOOKUP_PIN = os.environ.get("FUEL_LOOKUP_PIN", "1234")


def make_output_filename(original_filename, suffix=" extracted data", new_extension=".xlsx"):
    base_name = os.path.splitext(os.path.basename(original_filename))[0]
    return f"{base_name}{suffix}{new_extension}"


def data_path(filename):
    return os.path.join(os.path.dirname(__file__), "data", filename)


def save_uploaded_file(file_storage, suffix):
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix, prefix="upload_") as tmp:
        file_storage.save(tmp.name)
        return tmp.name


def save_base64_pdf(base64_text):
    if "," in base64_text:
        base64_text = base64_text.split(",", 1)[1]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf", prefix="upload_json_") as tmp:
        tmp.write(base64.b64decode(base64_text))
        return tmp.name


def money_ex_fuel(total_cost, fuel_percent):
    if fuel_percent <= 0:
        return float(Decimal(str(total_cost)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

    result = (
        Decimal(str(total_cost)) /
        (Decimal("1") + Decimal(str(fuel_percent)) / Decimal("100"))
    ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    return float(result)


def normalise_date_text(value):
    if value is None:
        return ""

    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")

    text = str(value).strip()
    if not text:
        return ""

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass

    return text


def parse_date_for_sort(text):
    if not text:
        return datetime.min

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(str(text).strip(), fmt)
        except ValueError:
            pass

    return datetime.min


def extract_connotes_from_pdf(filepath):
    rows = []

    with pdfplumber.open(filepath) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

    in_table = False

    for line in full_text.splitlines():
        if "Customer Con Note Sender Name Receiver Name" in line:
            in_table = True
            continue

        if in_table:
            if "Manifest Notes:" in line or "Total Connotes:" in line:
                break

            line = line.strip()
            if not line:
                continue

            parts = line.split()

            # Expected Northline row ending:
            # ... TotalItems Weight Cubic $TotalCost
            if len(parts) >= 5:
                connote = parts[1].strip()
                weight = parts[-3].replace(",", "").strip()
                raw_cubic = parts[-2].replace(",", "").strip()
                total_cost = parts[-1].replace("$", "").replace(",", "").strip()

                if (
                    re.fullmatch(r"[A-Z0-9\-]{6,30}", connote)
                    and re.fullmatch(r"\d+(\.\d+)?", weight)
                    and re.fullmatch(r"\d+(\.\d+)?", raw_cubic)
                    and re.fullmatch(r"\d+(\.\d+)?", total_cost)
                ):
                    rows.append({
                        "connote": connote,
                        "weight": float(weight),
                        "cubic": float(raw_cubic) * 250,
                        "total_cost": float(total_cost)
                    })

    return rows


def build_northline_workbook(all_rows, include_cost_report, fuel_percent):
    wb = Workbook()
    ws = wb.active
    ws.title = "Connotes"

    if include_cost_report:
        ws.append(["Connote Number", "Weight (Kg)", "Cubic Weight", "Cost Ex Fuel"])
    else:
        ws.append(["Connote Number", "Weight (Kg)", "Cubic Weight"])

    for row in all_rows:
        if include_cost_report:
            ws.append([
                row["connote"],
                row["weight"],
                row["cubic"],
                money_ex_fuel(row["total_cost"], fuel_percent),
            ])
        else:
            ws.append([row["connote"], row["weight"], row["cubic"]])

    for cell in ws["B"][1:]:
        cell.number_format = "0.00"

    for cell in ws["C"][1:]:
        cell.number_format = "0.00"

    if include_cost_report:
        for cell in ws["D"][1:]:
            cell.number_format = "$#,##0.00"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def safe_numeric_eval(formula_text):
    if not isinstance(formula_text, str):
        return formula_text

    text = formula_text.strip()
    if not text.startswith("="):
        return formula_text

    expr = text[1:].strip()
    if not re.fullmatch(r"[0-9\.\+\-\*/\(\)\s]+", expr):
        return formula_text

    try:
        return eval(expr, {"__builtins__": {}}, {})
    except Exception:
        return formula_text


def normalize_header(text):
    if text is None:
        return ""
    return str(text).strip().lower().replace(" ", "")


def find_column_indexes(header_row):
    headers = [normalize_header(x) for x in header_row]

    connote_idx = None
    amount_idx = None
    first_comment_idx = None

    for i, header in enumerate(headers):
        if connote_idx is None and header == "connotecode":
            connote_idx = i
        if amount_idx is None and header == "amount":
            amount_idx = i
        if first_comment_idx is None and header == "comment":
            first_comment_idx = i

    return connote_idx, amount_idx, first_comment_idx


def is_valid_connote_code(value):
    if value is None:
        return False

    text = str(value).strip()

    if not text:
        return False

    # Remove .0 if Excel stores whole-number connotes as numbers
    if text.endswith(".0"):
        text = text[:-2]

    # Allow letters, numbers, and hyphens
    if not re.fullmatch(r"[A-Z0-9\-]+", text, flags=re.IGNORECASE):
        return False

    # Must contain at least one number
    if not re.search(r"\d", text):
        return False

    # Must be long enough to be a real connote
    if len(text) < 6:
        return False

    return True


def extract_amount_rows_from_excel(filepath):
    extracted_rows = []

    wb_values = load_workbook(filepath, data_only=True)
    wb_formulas = load_workbook(filepath, data_only=False)

    for sheet_name in wb_values.sheetnames:
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]

        if ws_values.max_row < 2:
            continue

        header_row = [ws_values.cell(1, c).value for c in range(1, ws_values.max_column + 1)]
        connote_idx, amount_idx, first_comment_idx = find_column_indexes(header_row)

        if connote_idx is None or amount_idx is None or first_comment_idx is None:
            continue

        connote_col = connote_idx + 1
        amount_col = amount_idx + 1
        comment_col = first_comment_idx + 1

        for row_num in range(2, ws_values.max_row + 1):
            connote = ws_values.cell(row_num, connote_col).value
            amount = ws_values.cell(row_num, amount_col).value
            comment = ws_values.cell(row_num, comment_col).value

            if amount is None or str(amount).strip() == "":
                formula_value = ws_formulas.cell(row_num, amount_col).value
                amount = safe_numeric_eval(formula_value)

            if amount is None or str(amount).strip() == "":
                continue

            if not is_valid_connote_code(connote):
                continue

            connote_text = str(connote).strip()
            if connote_text.endswith(".0"):
                connote_text = connote_text[:-2]

            comment_text = "" if comment is None else str(comment).strip()
            extracted_rows.append([connote_text, amount, comment_text])

    return extracted_rows


def load_contacts():
    contacts = {}

    path = data_path("contacts_export.csv")
    if not os.path.exists(path):
        return contacts

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            company = (
                row.get("Company")
                or row.get("company")
                or row.get("COMPANY")
                or ""
            ).strip()

            levy_group = (
                row.get("Levy Group")
                or row.get("levy group")
                or row.get("LEVY GROUP")
                or ""
            ).strip()

            if company and levy_group:
                contacts[company] = levy_group

    return contacts


def load_levy_data(include_rates=False):
    """Load fuel history.

    Supports both formats:
    1. New/customer-specific format:
       Company, Levy Group, Effective Date, Fuel Surcharge
    2. Older/group-only format:
       Levy Group, Effective Date, Fuel Surcharge
    3. Original wide format:
       Effective Date, Run Date, Diesel Price, Customer A 1.15, Customer B 1.10...

    Returned dictionary keys are either Company names (for customer-specific rows)
    or Levy Group names (for group-level fallback rows).
    """
    levy_data = {}

    path = data_path("fuel_levy_history_report.csv")
    if not os.path.exists(path):
        return levy_data

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []

        has_company = "Company" in fieldnames
        has_clean_columns = "Levy Group" in fieldnames and "Effective Date" in fieldnames and "Fuel Surcharge" in fieldnames

        # New/customer-specific or old/group-only clean format
        if has_clean_columns:
            for row in reader:
                company = (row.get("Company") or "").strip()
                group = (row.get("Levy Group") or "").strip()
                date = normalise_date_text(row.get("Effective Date"))
                rate = (row.get("Fuel Surcharge") or "").strip()

                if not date or not rate:
                    continue

                # If Company exists, key by Company. Otherwise fall back to Levy Group.
                lookup_key = company if company else group
                if not lookup_key:
                    continue

                item = {"date": date}
                if include_rates:
                    item["rate"] = rate
                    item["levy_group"] = group

                levy_data.setdefault(lookup_key, []).append(item)

        # Original wide format: levy groups are column headers
        else:
            for row in reader:
                date = normalise_date_text(
                    row.get("Effective Date")
                    or row.get("effective date")
                    or row.get("Current Effective Date")
                    or ""
                )

                if not date:
                    continue

                for group, rate in row.items():
                    if group is None:
                        continue

                    group = group.strip()

                    if group in ["Effective Date", "Run Date", "Diesel Price (cpl)", "Status", "Emails Sent"]:
                        continue

                    if not group:
                        continue

                    rate_text = "" if rate is None else str(rate).strip()

                    if not rate_text:
                        continue

                    item = {"date": date}
                    if include_rates:
                        item["rate"] = rate_text
                        item["levy_group"] = group

                    levy_data.setdefault(group, []).append(item)

    # Deduplicate by date within each key, then sort newest first
    for key in list(levy_data.keys()):
        by_date = {}
        for item in levy_data[key]:
            by_date[item.get("date")] = item

        levy_data[key] = sorted(
            by_date.values(),
            key=lambda item: parse_date_for_sort(item.get("date")),
            reverse=True
        )

    return levy_data


def build_customer_lookup_data(include_rates=False):
    """Return lookup data keyed by customer name.

    Customer-specific rows are preferred. If a customer has no direct rows,
    it falls back to their Levy Group from contacts_export.csv.
    """
    contacts = load_contacts()
    raw_levy_data = load_levy_data(include_rates=include_rates)

    customer_lookup = {}
    customer_dates = {}

    for customer, group in contacts.items():
        direct_rows = raw_levy_data.get(customer, [])
        group_rows = raw_levy_data.get(group, [])

        combined = {}

        # Add group rows first, then customer rows override same-date values.
        for row in group_rows:
            combined[row.get("date")] = row

        for row in direct_rows:
            combined[row.get("date")] = row

        rows = sorted(
            combined.values(),
            key=lambda item: parse_date_for_sort(item.get("date")),
            reverse=True
        )

        customer_lookup[customer] = customer
        customer_dates[customer] = rows

    return customer_lookup, customer_dates


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/fuel-data", methods=["GET"])
def fuel_data():
    # Important: Do NOT send rates to the browser here.
    # The rates are only returned by /fuel-rate-secure after PIN check.
    contacts, customer_dates = build_customer_lookup_data(include_rates=False)
    return {
        "contacts": contacts,
        "levy_data": customer_dates,
    }


@app.route("/fuel-rate-secure", methods=["POST"])
def fuel_rate_secure():
    data = request.get_json(force=True)
    customer = data.get("customer")
    date = data.get("date")
    pin = data.get("pin")

    if pin != FUEL_LOOKUP_PIN:
        return {"ok": False, "error": "Invalid PIN"}, 403

    contacts, customer_dates = build_customer_lookup_data(include_rates=True)

    if customer not in contacts:
        return {"ok": False, "error": "Customer not found"}, 404

    for row in customer_dates.get(customer, []):
        if row.get("date") == date:
            return {"ok": True, "rate": row.get("rate", "")}

    return {"ok": False, "error": "Rate not found"}, 404


@app.route("/convert-fuel-levy", methods=["POST"])
def convert_fuel_levy():
    file = request.files.get("file")
    if not file:
        return "No file uploaded", 400

    existing_rows = {}
    existing_path = data_path("fuel_levy_history_report.csv")

    # 1. Load existing history first. Supports:
    #    Company, Levy Group, Effective Date, Fuel Surcharge
    #    Levy Group, Effective Date, Fuel Surcharge
    #    and the old wide format.
    if os.path.exists(existing_path):
        with open(existing_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames or []

            has_clean_columns = "Levy Group" in fieldnames and "Effective Date" in fieldnames and "Fuel Surcharge" in fieldnames

            if has_clean_columns:
                for row in reader:
                    company = (row.get("Company") or "").strip()
                    levy = (row.get("Levy Group") or "").strip()
                    date = normalise_date_text(row.get("Effective Date"))
                    rate = (row.get("Fuel Surcharge") or "").strip()

                    if levy and date and rate:
                        existing_rows[(company, levy, date)] = {
                            "Company": company,
                            "Levy Group": levy,
                            "Effective Date": date,
                            "Fuel Surcharge": rate,
                        }
            else:
                for row in reader:
                    date = normalise_date_text(row.get("Effective Date") or row.get("Current Effective Date") or "")

                    if not date:
                        continue

                    for group, rate in row.items():
                        if group is None:
                            continue

                        group = group.strip()

                        if group in ["Effective Date", "Run Date", "Diesel Price (cpl)", "Status", "Emails Sent"]:
                            continue

                        rate_text = "" if rate is None else str(rate).strip()

                        if group and rate_text:
                            existing_rows[("", group, date)] = {
                                "Company": "",
                                "Levy Group": group,
                                "Effective Date": date,
                                "Fuel Surcharge": rate_text,
                            }

    # 2. Load uploaded Excel file
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    header_row_num = None
    headers = []

    # Find real header row within first 20 rows
    for row_num in range(1, min(ws.max_row, 20) + 1):
        row_headers = []
        for cell in ws[row_num]:
            row_headers.append("" if cell.value is None else str(cell.value).strip().lower())

        if (
            "levy type" in row_headers
            and "customer / contractor" in row_headers
            and "current rate" in row_headers
            and "current effective date" in row_headers
        ):
            header_row_num = row_num
            headers = row_headers
            break

    if header_row_num is None:
        return "Could not find required columns: Levy Type, Customer / Contractor, Current Rate, Current Effective Date", 400

    levy_idx = headers.index("levy type")
    company_idx = headers.index("customer / contractor")
    rate_idx = headers.index("current rate")
    date_idx = headers.index("current effective date")

    # 3. Add new uploaded rows into existing rows
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        levy = row[levy_idx]
        company = row[company_idx]
        rate = row[rate_idx]
        date = row[date_idx]

        if levy is None or company is None or rate is None or date is None:
            continue

        levy_text = str(levy).strip()
        company_text = str(company).strip()
        rate_text = str(rate).strip()
        date_text = normalise_date_text(date)

        if not levy_text or not company_text or not rate_text or not date_text:
            continue

        # Key prevents duplicates for same customer/levy/date. New upload overwrites old row for same key.
        existing_rows[(company_text, levy_text, date_text)] = {
            "Company": company_text,
            "Levy Group": levy_text,
            "Effective Date": date_text,
            "Fuel Surcharge": rate_text,
        }

    sorted_rows = sorted(
        existing_rows.values(),
        key=lambda row: (
            parse_date_for_sort(row["Effective Date"]),
            row["Company"].lower(),
            row["Levy Group"].lower(),
        ),
        reverse=True
    )

    # 4. Output full replacement CSV
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Company", "Levy Group", "Effective Date", "Fuel Surcharge"])

    for row in sorted_rows:
        writer.writerow([
            row["Company"],
            row["Levy Group"],
            row["Effective Date"],
            row["Fuel Surcharge"],
        ])

    # IMPORTANT: this downloaded file is designed to replace data/fuel_levy_history_report.csv
    download_name = "fuel_levy_history_report.csv"

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f'attachment; filename="{download_name}"'
    response.headers["Content-Type"] = "text/csv; charset=utf-8"

    return response


@app.route("/extract-pdf-connotes-json", methods=["POST"])
def extract_pdf_connotes_json():
    try:
        payload = request.get_json(force=True)

        files = payload.get("files", [])
        remove_duplicates = bool(payload.get("remove_duplicates", False))
        include_cost_report = bool(payload.get("include_cost_report", False))

        try:
            fuel_percent = float(payload.get("fuel_surcharge_percent") or 0)
        except ValueError:
            fuel_percent = 0

        all_rows = []
        original_names = []

        for item in files:
            filename = item.get("name", "upload.pdf")
            content = item.get("content", "")

            if not filename.lower().endswith(".pdf"):
                continue

            original_names.append(filename)
            temp_path = save_base64_pdf(content)

            try:
                all_rows.extend(extract_connotes_from_pdf(temp_path))
            finally:
                try:
                    os.remove(temp_path)
                except Exception:
                    pass

        if remove_duplicates:
            seen = set()
            unique_rows = []
            for row in all_rows:
                if row["connote"] not in seen:
                    seen.add(row["connote"])
                    unique_rows.append(row)
            all_rows = unique_rows

        output = build_northline_workbook(all_rows, include_cost_report, fuel_percent)

        if len(original_names) == 1:
            download_name = make_output_filename(original_names[0])
        else:
            download_name = "combined extracted data.xlsx"

        response = make_response(send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ))

        response.headers["X-Extracted-Rows"] = str(len(all_rows))
        return response

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/extract-amount-comments", methods=["POST"])
def extract_amount_comments():
    files = request.files.getlist("excel_files")
    all_rows = []
    original_names = []

    for file in files:
        if file and file.filename and file.filename.lower().endswith((".xlsx", ".xlsm")):
            original_names.append(file.filename)
            suffix = os.path.splitext(file.filename)[1] or ".xlsx"
            temp_path = save_uploaded_file(file, suffix)

            try:
                all_rows.extend(extract_amount_rows_from_excel(temp_path))
            finally:
                try:
                    os.remove(temp_path)
                except Exception:
                    pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Amount Extract"
    ws.append(["Connote Code", "Amount", "First Comment"])

    for row in all_rows:
        ws.append(row)

    for cell in ws["B"][1:]:
        cell.number_format = "$#,##0.00"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    if len(original_names) == 1:
        download_name = make_output_filename(original_names[0])
    else:
        download_name = "combined extracted data.xlsx"

    response = make_response(send_file(
        output,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ))

    response.headers["X-Extracted-Rows"] = str(len(all_rows))
    return response


if __name__ == "__main__":
    app.run(debug=True)
